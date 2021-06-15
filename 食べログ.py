#!/usr/bin/env python
# coding: utf-8

# In[ ]:

# In[231]:


from bs4 import BeautifulSoup
import urllib.request as req
from selenium import webdriver
import time
import pandas as pd
#import chromedriver_binary
from selenium import webdriver
from selenium.webdriver.chrome.options import Options
from selenium.common.exceptions import NoSuchElementException
from selenium.common.exceptions import ElementNotInteractableException
from selenium.common.exceptions import TimeoutException
from selenium.common.exceptions import StaleElementReferenceException
from selenium.common.exceptions import ElementClickInterceptedException
import math
from selenium.webdriver.common.action_chains import ActionChains
from selenium.webdriver.common.keys import Keys
import pandas as pd
#import chromedriver_binary
import openpyxl
import glob
import xlrd
import pprint
from collections import OrderedDict
import csv
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import chromedriver_binary
from pyvirtualdisplay import Display
import re


# In[265]:


option = Options()# オプションを用意
#option.add_argument('--user-agent=hogehoge')
#option.add_argument('--headless')           # ヘッドレスモードの設定を付与
option.add_argument('--disable-extensions')       # すべての拡張機能を無効にする。ユーザースクリプトも無効にする
option.add_argument('--proxy-server="direct://"') # Proxy経由ではなく直接接続する
option.add_argument('--proxy-bypass-list=*')      # すべてのホスト名
option.add_argument('--blink-settings=imagesEnabled=false')#画像取得しない
option.add_argument("--start-maximized")
#option.add_argument('--start-maximized')          # 起動時にウィンドウを最大化する

#driver = webdriver.Chrome(options=option)   # Chromeを準備(optionでヘッドレスモードにしている）


# In[282]:


 
#driver = webdriver.Chrome(options=option) 
driver = webdriver.Chrome(executable_path='/Users/naoki/Downloads/chromedriver')
url_login=('https://tabelog.com/rstLst/?pcd=0&Cat=&RdoCosTp=2&LstCos=0&LstCosT=0&vac_net=0&search_date=2021%2F3%2F28%28%E6%97%A5%29&svt=1900&svps=2&svd=20210328&LstRev=0&sw=&award_prize%5B%5D=3&award_prize%5B%5D=2&award_prize%5B%5D=1&LstSitu=0&LstReserve=0&LstSmoking=0&PG=1&from_search=&voluntary_search=1&SrtT=trend&Srt=&sort_mode=&LstRange=&keyword=&from_search_form=1&lid=&ChkNewOpen=&hfc=1')
driver.get(url_login)
driver.maximize_window()


# In[280]:


namelists=[]
janlelists=[]
tellists=[]
address1lists=[]
address2lists=[]
address3lists=[]
timelists=[]
lunprilists=[]
dinprilists=[]
homepagelists=[]
sns1lists=[]
sns2lists=[]

i=0
while i!=1000:
        
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.CLASS_NAME, 'list-rst__rst-name-target')))
    browser_from=driver.find_elements_by_class_name('list-rst__rst-name-target')
    try: 
        #browser_from[i].click()
        webdriver.ActionChains(driver).move_to_element(browser_from[i]).perform()
   # except IndexError:
        ActionChains(driver).move_to_element(browser_from[i]).key_down(Keys.COMMAND).click().key_up(Keys.COMMAND).perform()
    except IndexError:
        
        nextpage=driver.find_elements_by_class_name('c-pagination__arrow')[-1]
        webdriver.ActionChains(driver).move_to_element(nextpage).perform()
        nextpage.click()
        i=0
        continue
    
    try:    
        handle_array = driver.window_handles
        driver.switch_to.window(handle_array[1])
    except IndexError:
        time.sleep(1)
        handle_array = driver.window_handles
        driver.switch_to.window(handle_array[1])
    name='-'
    janle='-'
    tel='-'
    address1='-'
    address2='-'
    address3='-'
    times='-'
    lunpri='-'
    dinpri='-'
    homepage='-'
    sns1='-'
    sns2='-'
    WebDriverWait(driver, 60).until(EC.presence_of_element_located((By.CLASS_NAME, 'gly-b-dinner')))
    try:
        name=driver.find_element_by_css_selector('#rst-data-head > table:nth-child(2) > tbody > tr:nth-child(1) > td').text
    except AttributeError:
        name='-'
    print(name)
    janle=driver.find_element_by_css_selector('#rst-data-head > table:nth-child(2) > tbody > tr:nth-child(3) > td').text
    print(janle)
    try:
        tel=driver.find_element_by_class_name('rstinfo-table__tel-num').text
    except NoSuchElementException:
        tel='-'
    print(tel)
    try:
        address1=driver.find_elements_by_class_name('listlink')[0].text
        print(address1)
        address2=driver.find_elements_by_class_name('listlink')[1].text
        print(address2)
        address3=driver.find_element_by_class_name('rstinfo-table__address').text
        print(address3)
    except IndexError:
        pass
    address3=address3.replace(address1,'')
    address3=address3.replace(address2,'')
    print(address3)
    try:
        times=driver.find_element_by_css_selector('#rst-data-head > table:nth-child(2) > tbody > tr:nth-child(7) > td > p:nth-child(2)').text
    except NoSuchElementException:
        pass
    if ('：' not in times)or(':' not in times):
        try:
            times=driver.find_element_by_css_selector('#rst-data-head > table:nth-child(2) > tbody > tr:nth-child(8) > td > p:nth-child(2)').text
        except NoSuchElementException:
            times='-'
    print(times)
    try:
        lunpri=driver.find_element_by_class_name('gly-b-lunch').text
    except NoSuchElementException:
        pass
    if lunpri=='-':
        try:
            lunpri=driver.find_elements_by_class_name('rstinfo-table__budget')[1].text
        except NoSuchElementException:
            lunpri='-'
            pass
        except IndexError: 
            lunpri='-'
            pass
    print(lunpri)
    try:
        dinpri=driver.find_element_by_class_name('gly-b-dinner').text
    except NoSuchElementException:
        dinpri='-'
    print(dinpri)
    try:
        homepage=driver.find_element_by_class_name('homepage').text
    except NoSuchElementException:
        homepage='-'
        print(homepage)
    try:
        sns1=driver.find_elements_by_class_name('rstinfo-sns-link')[0].text
    except IndexError:
        sns1='-'
        print(sns1)
        pass
    try:
        sns2=driver.find_elements_by_class_name('rstinfo-sns-link')[1].text
    except IndexError:
        sns2='-'
        print(sns2)
        pass
    namelists.append(name)
    janlelists.append(janle)
    tellists.append(tel)
    address1lists.append(address1)
    address2lists.append(address2)
    address3lists.append(address3)
    timelists.append(times)
    lunprilists.append(lunpri)
    dinprilists.append(dinpri)
    homepagelists.append(homepage)
    sns1lists.append(sns1)
    sns2lists.append(sns2)

    driver.close()
    handle_array = driver.window_handles
    driver.switch_to.window(handle_array[0])
    i=i+1


# In[281]:


#エクセルに出力
wb = openpyxl.load_workbook('/Users/naoki/Desktop/Mypandas/案件/食べログ/リストアップひな形.xlsx')
ws = wb['Sheet1']
for i in range(0,len(namelists)):
    ws.cell(row=i+1701,column=3,value=namelists[i])
    # A列
    ws.cell(row=i+1701,column=4,value=janlelists[i])
    # B列
    ws.cell(row=i+1701,column=5,value=tellists[i])
    ws.cell(row=i+1701,column=6,value=address1lists[i])
    ws.cell(row=i+1701,column=7,value=address2lists[i])
    ws.cell(row=i+1701,column=8,value=address3lists[i])
    ws.cell(row=i+1701,column=9,value=timelists[i])
    ws.cell(row=i+1701,column=11,value=lunprilists[i])
    ws.cell(row=i+1701,column=12,value=dinprilists[i])
    ws.cell(row=i+1701,column=13,value=homepagelists[i])
    ws.cell(row=i+1701,column=15,value=sns1lists[i])
    ws.cell(row=i+1701,column=16,value=sns2lists[i])
    
wb.save('/Users/naoki/Desktop/Mypandas/案件/食べログ/リストアップひな形.xlsx')






